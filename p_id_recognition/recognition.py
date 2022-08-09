
# built in
import os
import pathlib

# 3rd party
import numpy as np
import cv2
import pytesseract
import re
import xlwt
import openpyxl

'''
    -- show image
cv2.imshow (" circles ", img )
cv2.waitKey ()
'''


def recognize_instruments( pandidfname, noexcel, noannot):
    cur_dir = os.getcwd()
    img_fldr = os.path.join(cur_dir,"images/")
    circle_fldr = os.path.join(cur_dir,"images/", "circles/")
    stem_fname = pathlib.Path(pandidfname).stem
    
    # create xl
    wbcounter = 1
    pgcounter = 1
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='Instruments')
    sh = wb.get_sheet_by_name("Instruments")
    sh.cell(1,1).value = 'Instrument'
    sh.cell(1,2).value = 'Sheet'
    sh.cell(1,3).value = 'Page'
    sh.cell(1,4).value = 'x center point'
    sh.cell(1,5).value = 'y center point'
    sh.cell(1,6).value = 'r'
    
    # loop through each file in the img_fldr
    for img_path in [f for f in os.listdir(img_fldr) if os.path.isfile(os.path.join(img_fldr,f))]:
        
        # only look at files that are pngs
        if img_path.split ('.') [ -1] == 'png':
            # name of the image in the folder , e.g. 20
            img_name = img_path#.split ('.') [0]
            img_name_stem = pathlib.Path(img_name).stem
            #pgcounter += 1
            # Create txt and write headers
            file_path = os.path.join(circle_fldr+img_name_stem+".txt")
            file = open ( file_path , 'w')
            file.write ('instrument x y r:\n')
            img = cv2.imread ( img_fldr + img_path )
            gray = cv2.cvtColor ( img , cv2.COLOR_BGR2GRAY )
            
            
            # could include some image sizing selection to determine radius and houghcircle papameters
            # Hough Circle Transform with OpenCV
            circles1 = cv2.HoughCircles ( gray , cv2.HOUGH_GRADIENT , 1 , 10 , param1 =100 , 
                                          param2 =60 , minRadius =40, maxRadius =60) # worked for 2384x1684pdf
            #circles1 = cv2.HoughCircles ( gray , cv2.HOUGH_GRADIENT , 1 , 10 , param1 =100 , 
                                          #param2 =50 , minRadius =23, maxRadius =60) 
                                          # trying for 1191x841pdf -> 3573x2524png -> circle r=27
            
            
            try: # handle if there are not circles on the sheet
                circles = circles1 [0 , : , :]
            except: # continue to next image
                continue
            circles = np.uint16 ( np.around ( circles ) )
            counter = 0

            # draw on the detected circles
            for i in circles [:]:
                cv2.circle ( img , ( i [0] , i [1]) , i [2] , (255 , 0 , 255) , 7) # draw border
                cv2.circle ( img , ( i [0] , i [1]) , 3 , (255 , 0 , 255) , 7) # draw on center
            
            if not noannot:
                # save image
                cv2.imwrite (os.path.join(circle_fldr ,img_name_stem +'.png'), img )
            
            # get current sheet number
            sheet_n = recognize_sheet_numbers_for_sheet(img_fldr + img_path)
            
            # go through each circle that cv2 found
            for i in circles [:]:
                cv2.circle ( img , ( i [0] , i [1]) , i [2] , (255 , 255 , 255) , 7)

                # filter out the wrongly detected circles
                #if i [2] >= 25:
                if i [2] >= 25:
                    counter += 1
                    # generate the candidate area
                    #cropped_image_name = 'Dataset/images/circles/id_' + img_name + '/circle_' + str( counter ) + '.png'
                    cropped_circles = img [( i [1] - i [2]) :( i [1] + i [2]) , ( i [0] - i [2]) :( i [0] + i [2]) ]

                    # use Tesseract
                    cropped_circles_rgb = cv2.cvtColor ( cropped_circles ,cv2.COLOR_BGR2RGB )
                    # Recognition is very sensitive to this configuration
                    custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
                    text = pytesseract.image_to_string (cropped_circles_rgb , config = custom_config )
                    
                    # cv2.imshow (" circles ", cropped_circles_rgb )
                    # cv2.waitKey ()
                    # print("text",text)
                    
                    # Postprocessing Steps :
                    # 1. write in one line
                    text = text.replace ('\n', '').replace ('\r','').replace('|','')
                    # 2. Uppercase
                    text = text.upper ()
                    # 3. only keep the upper - case abbr and numbers
                    text = re.sub (u" ([^\ u0030 -\ u0039 \u0041 -\ u005a \ u002e])", "", text) # literall what
                    # 4. Some common problem ( specific to this dataset )
                    text = text.replace ('TL ', 'TI ')
                    text = text.replace ('POIE ', 'PDIE ')
                    
                    
                    # write to txt file
                    file.write ( text +' ') # instrument
                    file.write (str ( i [0]) +' ') # center point x
                    file.write (str ( i [1]) +' ') # center point y
                    file.write (str ( i [2]) +' ') #r
                    
                    file.write( sheet_n + "\n") # sheet number 
                    
                    # write to excel 
                    wbcounter = wbcounter + 1
                    if noexcel:
                        continue
                    sh.cell(wbcounter,1).value = text                       # tag
                    sh.cell(wbcounter,2).value = sheet_n                    # sheet
                    sh.cell(wbcounter,3).value = img_name_stem.split(" ")[-1]    # page
                    sh.cell(wbcounter,4).value = str ( i [0] )              # x
                    sh.cell(wbcounter,5).value = str ( i [1] )              # y
                    sh.cell(wbcounter,6).value = str ( i [2] )              # radius
            #file.save()
    if not noexcel:
        wb.save('outs\\Instrument Report for '+stem_fname+'.xlsx')
    
def recognize_sheet_numbers_for_sheet(original_image_name):
    ''' Takes an image object and returns the sheet number '''
    # gets passed a full path of an image to analyze
    # ie    C:/user/.../images/pdfname N.png
    
    original_image = cv2.imread ( original_image_name ) # create image object
    shape = original_image.shape
    h = shape [0] # Y ( height ) = 1786
    w = shape [1] # X ( width ) = 2526
    
    # crop image - these ratios work for my p&ids
    cropped_y_start = int( h * 0.97) # y start
    cropped_y_end = int( h * 1) # y end
    cropped_x_start = int( w * 0.94) # x start
    cropped_x_end = int( w * 1) # x end
    cropped_img = original_image [ cropped_y_start : cropped_y_end , 
                                   cropped_x_start : cropped_x_end ] # crop

    sheet_num = pytesseract.image_to_string ( cropped_img ) # get text from image
    sheet_num = sheet_num.replace (' ', '').replace ('SHEET','').replace('\n','') # replace dumb stuff
    
    return sheet_num
    
def recognize_sheet_numbers_for_document(pandidfname, noexcel):
    ''' Takes a link to the image folder, the document name, and a flag to 
        switch reporting to txt rather than excel.
        Uses Pytesseract to create a sheet to page number report '''
    img_fldr = os.path.join(os.getcwd(), "images")
    stem_fname = pathlib.Path(pandidfname).stem
    
    if not noexcel:
        # create excel doc with report
        wbcounter = 0
        wb = xlwt.Workbook()
        sheet1 = wb.add_sheet('Sheet to Page',cell_overwrite_ok=True)
        sheet1.write(wbcounter, 0, 'SHEET')
        sheet1.write(wbcounter, 1, 'PAGE')
    else:
        # create txt file to write report
        sheet2pagetxt = open ( os.path.join(os.getcwd(),"outs",
                               pandidfname + " - SHEET2PAGE.txt") , 'w')
        sheet2pagetxt.write("sheet\t\tpage\n")
    
    # go through files. assuming they all are pngs
    for file in [f for f in os.listdir(img_fldr) if os.path.isfile(os.path.join(img_fldr,f))]:
        
        # read in image and get shape
        original_image = cv2.imread ( os.path.join(img_fldr,file) )
        shape = original_image.shape
        h = shape [0] # Y ( height ) = 1786
        w = shape [1] # X ( width ) = 2526
        
        # crop image
        cropped_y_start = int( h * 0.97) # y start
        cropped_y_end = int( h * 1) # y end
        cropped_x_start = int( w * 0.94) # x start
        cropped_x_end = int( w * 1) # x end
        cropped_img = original_image [ cropped_y_start : cropped_y_end , 
                                       cropped_x_start : cropped_x_end ] # crop
        
        # get text on cropped image
        text = pytesseract.image_to_string ( cropped_img ) # get text from image
        text = text.replace (' ', '').replace ('SHEET','').replace('\n','') # replace dumb stuff
        
        # get image page number
        page_num = int(pathlib.Path(file).stem.split(" ")[-1])
        
        # write out depending on user selection
        if not noexcel:
            wbcounter += 1
            sheet1.write(wbcounter, 0, text) # sheet 
            sheet1.write(wbcounter, 1, page_num) # page
        else:
            sheet2pagetxt.write("{:<12s}{:<10s}\n".format(text, str(page_num)))
    
    # save excel doc
    if not noexcel:
        wb.save('outs\\Sheet to Page for '+stem_fname+'.xls')